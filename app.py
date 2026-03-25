import streamlit as st
import zipfile, io
from openpyxl import load_workbook, Workbook

st.set_page_config(page_title="Unificador de Hojas Excel", page_icon="📊")
st.title("📊 Unificador de Hojas Excel")
st.markdown("Subí tu archivo **.zip** con los archivos **.xlsx** y unificá todas las hojas en una sola por archivo.")

sep = st.radio("Separador entre hojas:", ["Sin separador", "Fila en blanco", "Fila con nombre de hoja"], horizontal=True)
skip_header = st.toggle("Encabezado una sola vez (omite el header en hojas 2, 3…)", value=True)

uploaded = st.file_uploader("Seleccioná tu archivo .zip", type="zip")

if uploaded:
    if st.button("⚡ Procesar ZIP", type="primary"):
        output_zip_buffer = io.BytesIO()
        with zipfile.ZipFile(io.BytesIO(uploaded.read()), 'r') as zin:
            nombres = [n for n in zin.namelist() if n.endswith(".xlsx") and "__MACOSX" not in n and not n.endswith("/")]
            st.info(f"📦 {len(nombres)} archivo(s) .xlsx encontrados.")
            if not nombres:
                st.error("No se encontraron archivos .xlsx en el ZIP.")
                st.stop()
            with zipfile.ZipFile(output_zip_buffer, 'w') as zout:
                bar = st.progress(0, text="Procesando...")
                for i, nombre in enumerate(nombres):
                    base = nombre.split("/")[-1]
                    st.write(f"▶ **{base}**")
                    try:
                        data = zin.read(nombre)
                        wb_in = load_workbook(io.BytesIO(data), data_only=True)
                        hojas = wb_in.sheetnames
                        if len(hojas) == 1:
                            st.write("  ✓ Una sola hoja, se copia sin cambios.")
                            zout.writestr(base, data)
                        else:
                            st.write(f"  • {len(hojas)} hoja(s): {', '.join(hojas)}")
                            wb_out = Workbook()
                            ws_out = wb_out.active
                            ws_out.title = "Unificado"
                            primera = True
                            for hoja in hojas:
                                ws_in = wb_in[hoja]
                                filas = list(ws_in.iter_rows(values_only=True))
                                if not filas:
                                    st.write(f"  ⚠️ Hoja '{hoja}' vacía, omitida.")
                                    continue
                                if not primera:
                                    if sep == "Fila en blanco":
                                        ws_out.append([])
                                    elif sep == "Fila con nombre de hoja":
                                        ws_out.append([f"▼ {hoja}"])
                                inicio = 1 if (not primera and skip_header and len(filas) > 1) else 0
                                for fila in filas[inicio:]:
                                    ws_out.append(list(fila))
                                st.write(f"  ✅ '{hoja}' → {len(filas) - inicio} fila(s) copiadas.")
                                primera = False
                            buf = io.BytesIO()
                            wb_out.save(buf)
                            zout.writestr(base, buf.getvalue())
                    except Exception as e:
                        st.error(f"❌ Error en '{base}': {e}")
                    bar.progress((i + 1) / len(nombres), text=f"Procesando {i+1}/{len(nombres)}...")
        st.success("🎉 ¡Listo! Descargá tu ZIP unificado:")
        st.download_button(label="⬇️ Descargar ZIP unificado", data=output_zip_buffer.getvalue(), file_name="unificado_" + uploaded.name, mime="application/zip")
